#!/usr/bin/python3
# -*- coding: UTF-8 -*-
__author__ = "一千零一点(崔嘉铭)"
__file__ = "main.py"
__time__ = "2024/6/10 22:11"

from openpyxl import load_workbook
from function import *
# 打开表格
work = load_workbook(filename="achievement.xlsx")
# 创造折线图的横坐标(考试名字)
rows = []
# 纵坐标(数据:分数)
data = {}
for i in range(7):
    data[i] = []
# print(data)
# print(data[2])
# 科目表
subject = []
# 打开对应sheet表
sheet = work["Sheet1"]
# 锁定第一个考试名字位置(高一第一学期月考)
row=2
column=1
# 获取所有考试名字
while True:
    Name = sheet.cell(row=row, column=column).value
    if Name is not None:
        rows.append(Name)
        row += 1
    else:
        # 锁定第一个科目(语文的)位置,归零化处理
        row=1
        column=2
        function1(rows=rows)
        break
# print(subject)
i=0
while True:
    Name = sheet.cell(row=row, column=column).value
    if Name is not None:
        if isinstance(Name, str):
            subject=Name
            # 打印科目
            # print(subject)
            row = row+1
        if isinstance(Name, int):
            # 打印分数
            # print(sheet.cell(row=row, column=column).value)
            data[i].append(Name)
            row = row+1
        if isinstance(Name, float):
            # 打印分数
            # print(sheet.cell(row=row, column=column).value)
            data[i].append(Name)
            row = row+1
    else:
        function2(subject=subject, data=data[i])
        # 锁定下一个学科
        column += 3
        row = 1
        i +=1
        if sheet.cell(row=row, column=column).value is None:
            break

creat()

